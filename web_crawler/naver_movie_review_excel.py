import requests
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook

url = "https://movie.naver.com/movie/bi/mi/pointWriteFormList.nhn?code=188909&type=after&isActualPointWriteExecute" \
      "=false&isMileageSubscriptionAlready=false&isMileageSubscriptionReject=false"

html = bs(requests.get(url).content, "lxml", from_encoding="utf-8")
count = html.select("body > div > div > div.score_total > strong > em")[0].contents[0].replace(',', '')
print(count)

wb = Workbook()
ws = wb.create_sheet("겨울왕국2", 0)
row = 2
ws.cell(1,1, "평점")
ws.cell(1,2, "좋아요")
ws.cell(1,3, "싫어요")
# ws.cell(1,4, "비율")
ws.cell(1,4, "댓글")

# score = html.select("body > div > div > div.score_result > ul > li > div.star_score > em")
# reple = html.select("body > div > div > div.score_result > ul > li > div.score_reple > p")
# like = html.select("body > div > div > div.score_result > ul > li > div.btn_area")
# print(reple)
# for x in range(1, int(count)//10 + 1):
score_li = []

for x in range(1, 6):
    html = bs(requests.get(url + "&page=" + str(x) ).content, "html.parser", from_encoding="utf-8")
    scores = html.select("body > div > div > div.score_result > ul > li > div.star_score > em")
    reples = html.select("body > div > div > div.score_result > ul > li > div.score_reple > p")
    goods = html.select("body > div > div > div.score_result > ul > li > div.btn_area")
    # print(reple)
    for i in range(len(reples)):

        score = scores[i].contents[0]
        good = goods[i].contents[1].contents[5].contents[0]
        bad = goods[i].contents[3].contents[5].contents[0]
        reple = ""
        # print(reples[i].contents)

        # if reples[i].contents[1] ==" 스포일러 컨텐츠로 처리되는지 여부 ":
        #     reple = reples[i].contents[3].contents[0].strip()
        # else:
        #     reple = reples[i].contents[5].contents[0].strip()

        try :
            reple = reples[i].contents[5].contents[0].strip()
        except:
            reple = reples[i].contents[3].contents[0].strip()

        if reple=="":
            try:
                reple = reples[i].contents[5].contents[1].contents[1]['data-src']
            except:
                continue

        print(i, "=" * 50)
        print("평점 :", score)
        print("좋아요 :", good)
        print("싫어요 :", bad)
        print("댓글 :", reple)

        ws.cell(row, 1, int(score))
        ws.cell(row, 2, int(good))
        ws.cell(row, 3, int(bad))
        ws.cell(row, 4, reple)
        row += 1

wb.save("강철비2.xlsx")
print("=" * 50)
print("데이터가 저장 되었습니다.")
print("=" * 50)
